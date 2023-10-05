import datetime
import os
from openpyxl import load_workbook
from openpyxl import Workbook

currDate = datetime.date.today()
cDate = currDate.strftime("%d %b %Y").upper()
print("Current Date: ", cDate)

# To run type command in Terminal: python.exe C:\PyCharm_DCO\readExcel\tests\home\mySandbox.py
# Console will require to change getcwd()
class Sandbox:

    def sheetData(self, sheet):
        maxCol = sheet.max_column
        userData = {}
        columns = []
        for col in sheet.iter_cols(min_row=1,
                                   min_col=1,
                                   max_col=maxCol,
                                   values_only=True):
            columnName = col[0]
            columns.append(columnName)
        for row in sheet.iter_rows(min_row=2,
                                   min_col=1,
                                   max_col=maxCol,
                                   values_only=True):
            product_id = row[0]
            product = {
                columns[1]: row[1],
                columns[2]: row[2],
                columns[3]: row[3],
                columns[4]: row[4],
                columns[5]: row[5],
                columns[6]: row[6],
                columns[7]: row[7],
                columns[8]: row[8],
                columns[9]: row[9],
                columns[10]: row[10],
                columns[11]: row[11],
                columns[12]: row[12],
                columns[13]: row[13],
                columns[14]: row[14],
                columns[15]: row[15],
                columns[16]: row[16],
                columns[17]: row[17],
                columns[18]: row[18],
                columns[19]: row[19],
                columns[20]: row[20],
                columns[21]: row[21],
                columns[22]: row[22],
                columns[23]: row[23],
                columns[24]: row[24],
                columns[25]: row[25],
                columns[26]: row[26],
                columns[27]: row[27],
                columns[28]: row[28],
                columns[29]: row[29],
                columns[30]: row[30],
                columns[31]: row[31],
                columns[32]: row[32],
                columns[33]: row[33]
            }
            userData[product_id] = product
        return userData

    def openInput(self):
        dirPath = os.getcwd()
        iFilepath = f"{dirPath}\\Data.xlsx"
        print('Data file: ' + iFilepath)
        iWB = load_workbook(iFilepath)
        iSheet = iWB["MDCO-E2E"]
        return iSheet

    def openOutput(self):
        dirPath = os.getcwd()
        oFilepath = f"{dirPath}\\testData.xlsx"
        try:
            oWB = load_workbook(oFilepath)
            oSheet = oWB["Sheet1"]
        except FileNotFoundError:
            oWB = Workbook()
            oSheet = oWB.active
        return [oFilepath, oWB, oSheet]


if __name__ == '__main__':
    sb = Sandbox()
    inFile = sb.openInput()
    output = sb.openOutput()
    filepath = output[0]
    wb = output[1]
    sheet = output[2]
    excelSheetData = sb.sheetData(inFile)
    ctr = 1
    for e in excelSheetData:
        print(e)
        output.append(e)
        ctr += 1
        sheet.cell(row=ctr, column=1).value = e
    wb.save(filepath)

